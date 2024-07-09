import io
import pandas
import requests
import hashlib
import openpyxl
import numpy as np
import pandas as pd
import pydantic


class TableHead(pydantic.BaseModel):
    name: str
    time: str
    budget_places: int
    paid_places: int


def download_file(url):
    return requests.get(url, allow_redirects=True).content


def write_file(content: bytes, path):
    with open(path, 'wb') as file:
        file.write(content)


def md5(content: bytes):
    return hashlib.md5(content).hexdigest()


def get_table_head(file: str | io.BytesIO):
    wb = openpyxl.load_workbook(file)
    ws = wb.active
    name = ws['A2'].value
    time = ws['F5'].value
    budget_places = ws['F8'].value
    paid_places = ws['F10'].value
    return TableHead(name=name, time=time, budget_places=budget_places, paid_places=paid_places)


def parse_xlsx(file: str | io.BytesIO):
    table_head = get_table_head(file)
    print(table_head)

    df = pd.read_excel(file, index_col=0, header=14, engine='openpyxl')
    df: pandas.DataFrame = df.loc[:, ~df.columns.str.contains('^Unnamed')]
    cols = df.columns
    # df = df.sort_values(by="Сумма конкурсных баллов", ascending=False)
    print(df.shape[0])
    df_first_priority: pandas.DataFrame = df[df['Приоритет иных мест'] == 1]
    df_first_priority_original = df_first_priority[df_first_priority['Оригинал аттестата'] == 'Да']
    last = df_first_priority.iloc[table_head.budget_places - 1]
    print(df_first_priority_original.shape[0])
    print(last['Сумма конкурсных баллов'])


def main():
    content = download_file('https://enrol.hse.ru/storage/public_report_2024/moscow/Bachelors/BD_moscow_AM_O.xlsx')
    md5_hash = md5(content)
    print(md5_hash)
    parse_xlsx(io.BytesIO(content))


if __name__ == '__main__':
    main()
