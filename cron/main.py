import io
import json
import sys
import time
import schedule
import pandas
import openpyxl
import pandas as pd
import pydantic

from lib.utils import *
from pprint import pprint
from lib.api import get_directions, add_stats, send_push_notifications
from lib.init import update_timeout
from lib.logger import logger


class TableHead(pydantic.BaseModel):
    name: str
    time: str
    budget_places: int
    paid_places: int


def get_table_head(file: io.BytesIO):
    wb = openpyxl.load_workbook(file)
    ws = wb.active
    name = ws['A2'].value
    time = ws['F5'].value
    budget_places = ws['F8'].value
    paid_places = ws['F10'].value
    return TableHead(name=name, time=time, budget_places=budget_places, paid_places=paid_places)


def find_stat_by_text(text, previous_stats):
    return list(filter(lambda stat: stat['text'] == text, previous_stats))


def is_numeric(value):
    return isinstance(value, int) or (isinstance(value, str) and value.isnumeric())


def decorator_add_stat_val(previous_stats):
    def get_diff(text, value, previous_stats):
        if not previous_stats:
            return 0
        selected_data = find_stat_by_text(text, previous_stats)
        if len(selected_data) == 0:
            return 0
        previous_value = selected_data[0]['value']
        if not (is_numeric(previous_value) and is_numeric(value)):
            return 0
        return int(value) - int(previous_value)

    def wrapper(values: [], text: str, value: str | int):
        values.append({'text': text, 'value': str(value), 'diff': get_diff(text, value, previous_stats)})

    return wrapper


def print_stat(string: str, data):
    print(f'{string:<27}{data}')


def get_table_stats(table_head):
    return {'type': 'stats', 'head': 'Информация о направлении', 'values': [
        {'text': 'Бюджетные места', 'value': str(table_head.budget_places), 'diff': 0},
        {'text': 'Платные места', 'value': str(table_head.paid_places), 'diff': 0},
    ]}


def get_general_stats(df: pandas.DataFrame, previous_stats=None):
    values = []
    add_stat_val = decorator_add_stat_val(previous_stats)

    all_applications = df.shape[0]
    add_stat_val(values, "Количество заявлений:", all_applications)

    my_place = df.index[df['СНИЛС / Уникальный идентификатор'] == '173-102-564 30'].tolist()
    add_stat_val(values, "Мое место:", 'Сюда не подавал' if len(my_place) == 0 else my_place[0])

    target = df[df['Поступление на места в рамках квоты\nцелевого приема'] == 'Да'].shape[0]
    add_stat_val(values, "Целевое:", target)

    wee = df[df['Право поступления\nбез вступительных испытаний'] == 'Да'].shape[0]
    add_stat_val(values, "БВИ:", wee)

    separate_quota = df[df['Поступление на места\nв рамках отдельной квоты'] == 'Да'].shape[0]
    add_stat_val(values, "Отдельная квота:", separate_quota)

    special_right = df[df['Поступление на места в рамках квоты \nдля лиц, имеющих особое право'] == 'Да'].shape[0]
    add_stat_val(values, "Особое право:", special_right)

    original = df[df['Оригинал аттестата'] == 'Да'].shape[0]
    add_stat_val(values, "Оригинал аттестата:", original)

    if 'Высший приоритет' in df:
        first_high_priority = df[df['Высший приоритет'] == 1].shape[0]
    else:
        first_high_priority = 'Нет такой колонки!'

    add_stat_val(values, "Первый высший приоритет:", first_high_priority)

    return {'type': 'stats', 'head': 'Общая статистика', 'values': values}


def get_first_priority_stats(df_general: pandas.DataFrame, budget_places, previous_stats=None):
    df = df_general[df_general['Приоритет иных мест'] == 1]
    values = []
    add_stat_val = decorator_add_stat_val(previous_stats)

    all_applications = df.shape[0]
    add_stat_val(values, "Количество заявлений:", all_applications)

    my_place = df.index[df['СНИЛС / Уникальный идентификатор'] == '173-102-564 30'].tolist()
    add_stat_val(values, "Мое место:", 'Сюда не подавал' if len(my_place) == 0 else my_place[0])

    target = df[df['Поступление на места в рамках квоты\nцелевого приема'] == 'Да'].shape[0]
    add_stat_val(values, "Целевое:", target)

    wee = df[df['Право поступления\nбез вступительных испытаний'] == 'Да'].shape[0]
    add_stat_val(values, "БВИ:", wee)

    separate_quota = df[df['Поступление на места\nв рамках отдельной квоты'] == 'Да'].shape[0]
    add_stat_val(values, "Отдельная квота:", separate_quota)

    special_right = df[df['Поступление на места в рамках квоты \nдля лиц, имеющих особое право'] == 'Да'].shape[0]
    add_stat_val(values, "Особое право:", special_right)

    original = df[df['Оригинал аттестата'] == 'Да'].shape[0]
    add_stat_val(values, "Оригинал аттестата:", original)

    if 'Высший приоритет' in df:
        first_high_priority = df[df['Высший приоритет'] == 1].shape[0]
    else:
        first_high_priority = 'Нет такой колонки!'

    add_stat_val(values, "Первый высший приоритет:", first_high_priority)

    last_budget = df.iloc[budget_places - 1]
    if (last_budget['Поступление на места в рамках квоты\nцелевого приема'] == "Да") or \
            (last_budget['Право поступления\nбез вступительных испытаний'] == "Да") or \
            (last_budget['Поступление на места\nв рамках отдельной квоты'] == "Да") or \
            (last_budget['Поступление на места в рамках квоты \nдля лиц, имеющих особое право'] == "Да"):
        last_budget_mark = "Занято абитуриентами в приоритетном порядке"
    else:
        last_budget_mark = int(float(last_budget["Сумма конкурсных баллов"]))
    add_stat_val(values, "Проходной балл на бюджет:", last_budget_mark)

    last_budget_number = last_budget.name
    add_stat_val(values, "Номер крайнего бюджетника:", last_budget_number)

    budget = df[df['Вид места'] == 'Б'].shape[0]
    add_stat_val(values, "Бюджет:", budget)

    paid = df[df['Вид места'] == 'К'].shape[0]
    add_stat_val(values, "Коммерция:", paid)

    budget_paid = df[(df['Вид места'] == 'Б; К') | (df['Вид места'] == 'К; Б')].shape[0]
    add_stat_val(values, "Бюджет и коммерция:", budget_paid)

    return {'type': 'stats', 'head': 'Первый приоритет', 'values': values}


def parse_xlsx(file: io.BytesIO, previous_stats):
    stats = []

    table_head = get_table_head(file)
    df = pd.read_excel(file, index_col=0, header=14, engine='openpyxl')
    df = df.loc[:, ~df.columns.str.contains('^Unnamed')]

    # print(df.columns)
    stats.append(get_table_stats(table_head))
    if previous_stats:
        stats.append(get_general_stats(df, previous_stats[1]['values']))
        stats.append(get_first_priority_stats(df, table_head.budget_places, previous_stats[2]['values']))
    else:
        stats.append(get_general_stats(df))
        stats.append(get_first_priority_stats(df, table_head.budget_places))
    return table_head, stats


def update_hse_data(directions):
    logger.info('Updating hse information...')
    am_stats = ''
    for direction in directions:
        content = download_file(direction['url'])
        if not content:
            continue

        logger.info(f'File downloaded, size: {sys.getsizeof(content)} bytes')
        # pprint(parse_xlsx(io.BytesIO(content), direction["stats"])[1])
        md5_hash = md5(content)
        if md5_hash == direction['hash']:
            logger.info(f'Information about program "{direction["name"]}" not updated (hashsum not changed)')
            continue

        table_head, stats = parse_xlsx(io.BytesIO(content), direction["stats"])
        logger.info(f'Information about program "{direction["name"]}" successfully updated')

        stats_string = json.dumps(stats, ensure_ascii=False)
        logger.debug(f'Stats string length: {len(stats_string)}')
        add_stats(table_head.time, direction["name"], stats_string, md5_hash)
        logger.debug(f'Add stats to server')
        if direction["name"] == 'Прикладная математика':
            am_stats = stats

    if am_stats:
        wee = find_stat_by_text("БВИ:", am_stats[1]["values"])[0]["value"]
        my_place = find_stat_by_text("Мое место:", am_stats[1]["values"])[0]["value"]
        send_push_notifications(f'БВИ {wee}', f'Место {my_place}')
    logger.info('End updating hse information.')


def schedule_parsing():
    cur_time = get_time()
    update_interval = update_timeout
    for sensitive_time in [('11:00', '11:30'), ('13:00', '13:30'), ('15:00', '15:30'), ('17:00', '17:30'),
                           ('19:00', '19:30')]:
        if time_in_range(*sensitive_time, cur_time):
            update_interval = update_timeout // 5
            logger.debug(f'Sensitive time, set update interval to "{update_interval}"')
            break

    job = schedule.every(update_interval).seconds.do(update_parsing)
    logger.info(f'Schedule parsing (next_run: {job.next_run})')


def update_parsing():
    directions = get_directions()
    if not directions:
        logger.warning('No directions found!')
        schedule_parsing()
        return schedule.CancelJob

    update_hse_data(directions)

    schedule_parsing()
    return schedule.CancelJob


def main():
    update_parsing()
    while True:
        schedule.run_pending()
        time.sleep(1)


if __name__ == '__main__':
    try:
        main()
    except KeyboardInterrupt:
        logger.info('Stop cron process.')
